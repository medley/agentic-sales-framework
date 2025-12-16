#!/usr/bin/env python3
"""
Create ROI Template Excel File
Generates roi_template.xlsx with formulas, color coding, and proper structure
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_roi_template(output_path):
    """Create ROI template with 5 tabs and proper formulas"""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Color definitions
    BLUE_FILL = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')
    BLUE_FONT = Font(color='0000FF', bold=True)
    GREEN_FONT = Font(color='008000')
    BLACK_FONT = Font(bold=True)
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True, size=12)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== TAB 1: EXECUTIVE SUMMARY =====
    ws_exec = wb.create_sheet('Executive Summary', 0)
    ws_exec.column_dimensions['A'].width = 25
    ws_exec.column_dimensions['B'].width = 18

    # Header
    ws_exec['A1'] = 'ROI Executive Summary'
    ws_exec['A1'].font = Font(size=16, bold=True)
    ws_exec.merge_cells('A1:B1')

    # Data rows
    headers = [
        ('Investment (Year 1)', '=\'Financial Summary\'!B2', 'dollars'),
        ('Annual Value', '=\'Financial Summary\'!B3', 'dollars'),
        ('Payback Period (months)', '=\'Financial Summary\'!B4', 'number'),
        ('3-Year Net Value', '=\'Financial Summary\'!B5', 'dollars'),
        ('3-Year ROI Multiple', '=\'Financial Summary\'!B6', 'ratio')
    ]

    row = 3
    for label, formula, fmt in headers:
        ws_exec[f'A{row}'] = label
        ws_exec[f'A{row}'].font = Font(bold=True)
        ws_exec[f'B{row}'] = formula
        ws_exec[f'B{row}'].font = GREEN_FONT  # Cross-sheet reference

        if fmt == 'dollars':
            ws_exec[f'B{row}'].number_format = '$#,##0'
        elif fmt == 'number':
            ws_exec[f'B{row}'].number_format = '0.0'
        elif fmt == 'ratio':
            ws_exec[f'B{row}'].number_format = '0.0":1"'

        ws_exec[f'B{row}'].border = thin_border
        row += 1

    # ===== TAB 2: CURRENT STATE COSTS =====
    ws_current = wb.create_sheet('Current State Costs', 1)
    ws_current.column_dimensions['A'].width = 30
    ws_current.column_dimensions['B'].width = 18
    ws_current.column_dimensions['C'].width = 40

    # Header
    ws_current['A1'] = 'Current State Annual Costs'
    ws_current['A1'].font = Font(size=14, bold=True)
    ws_current.merge_cells('A1:C1')

    ws_current['A2'] = 'Cost Category'
    ws_current['B2'] = 'Annual Cost'
    ws_current['C2'] = 'Source / Notes'
    for col in ['A', 'B', 'C']:
        ws_current[f'{col}2'].fill = HEADER_FILL
        ws_current[f'{col}2'].font = HEADER_FONT

    # Cost categories (blue = input cells)
    costs = [
        ('QA Labor Inefficiency', 250000, 'FTE count × avg salary × % time on manual review'),
        ('Documentation Error Cost', 81000, 'Error rate × batch volume × rework cost'),
        ('Deviation Administrative Burden', 70000, 'Deviation count × hours per deviation × hourly rate'),
        ('Audit Preparation Time', 78000, 'Audit count × prep hours × hourly rate'),
        ('Compliance Risk Exposure', 200000, 'Potential FDA warning letter, customer fines, shipment holds')
    ]

    row = 3
    for label, default_value, note in costs:
        ws_current[f'A{row}'] = label
        ws_current[f'B{row}'] = default_value
        ws_current[f'B{row}'].fill = BLUE_FILL  # Input cell
        ws_current[f'B{row}'].font = BLUE_FONT
        ws_current[f'B{row}'].number_format = '$#,##0'
        ws_current[f'B{row}'].border = thin_border
        ws_current[f'C{row}'] = note
        ws_current[f'C{row}'].font = Font(italic=True, size=9)
        row += 1

    # Total row
    ws_current[f'A{row}'] = 'TOTAL ANNUAL COST'
    ws_current[f'A{row}'].font = Font(bold=True)
    ws_current[f'B{row}'] = f'=SUM(B3:B{row-1})'
    ws_current[f'B{row}'].font = BLACK_FONT  # Formula
    ws_current[f'B{row}'].number_format = '$#,##0'
    ws_current[f'B{row}'].border = Border(
        top=Side(style='double'),
        bottom=Side(style='double'),
        left=Side(style='thin'),
        right=Side(style='thin')
    )

    # ===== TAB 3: SOLUTION VALUE =====
    ws_solution = wb.create_sheet('Solution Value', 2)
    ws_solution.column_dimensions['A'].width = 35
    ws_solution.column_dimensions['B'].width = 18
    ws_solution.column_dimensions['C'].width = 50

    # Header
    ws_solution['A1'] = 'Annual Value Creation (Savings)'
    ws_solution['A1'].font = Font(size=14, bold=True)
    ws_solution.merge_cells('A1:C1')

    ws_solution['A2'] = 'Value Category'
    ws_solution['B2'] = 'Annual Savings'
    ws_solution['C2'] = 'Calculation Basis'
    for col in ['A', 'B', 'C']:
        ws_solution[f'{col}2'].fill = HEADER_FILL
        ws_solution[f'{col}2'].font = HEADER_FONT

    # Value categories (formulas based on Current State)
    values = [
        ('QA Efficiency (80% time savings)', '=\'Current State Costs\'!B3*0.80', '80% reduction in QA review time'),
        ('Error Elimination (90% reduction)', '=\'Current State Costs\'!B4*0.90', 'Eliminate documentation errors via digital capture'),
        ('Deviation Reduction (80% reduction)', '=\'Current State Costs\'!B5*0.80', 'Alternate BOM eliminates planned deviations'),
        ('Audit Efficiency (50% faster prep)', '=\'Current State Costs\'!B6*0.50', 'Electronic records = faster audit prep'),
        ('Vendor Consolidation', 40000, 'Replace existing QMS/complaints tool (input if applicable)')
    ]

    row = 3
    for label, formula_or_value, note in values:
        ws_solution[f'A{row}'] = label
        if isinstance(formula_or_value, str) and formula_or_value.startswith('='):
            ws_solution[f'B{row}'] = formula_or_value
            ws_solution[f'B{row}'].font = GREEN_FONT  # Cross-sheet formula
        else:
            ws_solution[f'B{row}'] = formula_or_value
            ws_solution[f'B{row}'].fill = BLUE_FILL  # Input cell
            ws_solution[f'B{row}'].font = BLUE_FONT
        ws_solution[f'B{row}'].number_format = '$#,##0'
        ws_solution[f'B{row}'].border = thin_border
        ws_solution[f'C{row}'] = note
        ws_solution[f'C{row}'].font = Font(italic=True, size=9)
        row += 1

    # Total row
    ws_solution[f'A{row}'] = 'TOTAL ANNUAL VALUE'
    ws_solution[f'A{row}'].font = Font(bold=True)
    ws_solution[f'B{row}'] = f'=SUM(B3:B{row-1})'
    ws_solution[f'B{row}'].font = BLACK_FONT  # Formula
    ws_solution[f'B{row}'].number_format = '$#,##0'
    ws_solution[f'B{row}'].border = Border(
        top=Side(style='double'),
        bottom=Side(style='double'),
        left=Side(style='thin'),
        right=Side(style='thin')
    )

    # ===== TAB 4: FINANCIAL SUMMARY =====
    ws_financial = wb.create_sheet('Financial Summary', 3)
    ws_financial.column_dimensions['A'].width = 30
    ws_financial.column_dimensions['B'].width = 18
    ws_financial.column_dimensions['C'].width = 50

    # Header
    ws_financial['A1'] = 'Financial Summary & ROI'
    ws_financial['A1'].font = Font(size=14, bold=True)
    ws_financial.merge_cells('A1:C1')

    # Data rows
    financial_data = [
        ('Investment (Year 1)', 220000, 'ACV + implementation services (~$50-75K)', 'blue'),
        ('Annual Value', '=\'Solution Value\'!B7', 'Total annual savings from solution', 'green'),
        ('Payback Period (months)', '=B2/(B3/12)', 'Investment ÷ (Annual Value / 12)', 'black'),
        ('3-Year Net Value', '=(B3*3)-B2', '(Annual Value × 3) - Investment', 'black'),
        ('3-Year ROI Multiple', '=B5/B2', '3-Year Net Value ÷ Investment', 'black')
    ]

    row = 2
    for label, formula_or_value, note, color in financial_data:
        ws_financial[f'A{row}'] = label
        ws_financial[f'A{row}'].font = Font(bold=True)

        if isinstance(formula_or_value, str) and formula_or_value.startswith('='):
            ws_financial[f'B{row}'] = formula_or_value
            if color == 'green':
                ws_financial[f'B{row}'].font = GREEN_FONT
            else:
                ws_financial[f'B{row}'].font = BLACK_FONT
        else:
            ws_financial[f'B{row}'] = formula_or_value
            ws_financial[f'B{row}'].fill = BLUE_FILL
            ws_financial[f'B{row}'].font = BLUE_FONT

        # Number formatting
        if 'Investment' in label or 'Value' in label:
            ws_financial[f'B{row}'].number_format = '$#,##0'
        elif 'Payback' in label:
            ws_financial[f'B{row}'].number_format = '0.0'
        elif 'ROI' in label:
            ws_financial[f'B{row}'].number_format = '0.0":1"'

        ws_financial[f'B{row}'].border = thin_border
        ws_financial[f'C{row}'] = note
        ws_financial[f'C{row}'].font = Font(italic=True, size=9)
        row += 1

    # ===== TAB 5: ASSUMPTIONS =====
    ws_assumptions = wb.create_sheet('Assumptions', 4)
    ws_assumptions.column_dimensions['A'].width = 35
    ws_assumptions.column_dimensions['B'].width = 50

    # Header
    ws_assumptions['A1'] = 'Model Assumptions & Inputs'
    ws_assumptions['A1'].font = Font(size=14, bold=True)
    ws_assumptions.merge_cells('A1:B1')

    ws_assumptions['A3'] = 'Input Parameter'
    ws_assumptions['B3'] = 'Source / Justification'
    for col in ['A', 'B']:
        ws_assumptions[f'{col}3'].fill = HEADER_FILL
        ws_assumptions[f'{col}3'].font = HEADER_FONT

    # Assumptions list
    assumptions = [
        ('Conservative Estimates Used', 'All savings estimates are conservative (e.g., 80% vs potential 90%+ savings)'),
        ('QA Time Savings: 80%', 'Review-by-exception vs line-by-line review (industry standard: 70-90%)'),
        ('Error Reduction: 90%', 'Digital capture eliminates legibility, calculation, missing entry errors'),
        ('Deviation Reduction: 80%', 'Alternate BOM capability eliminates planned deviations'),
        ('Audit Efficiency: 50%', 'Electronic records reduce audit prep time by half'),
        ('Average QA Reviewer Salary: $100K', 'Industry standard for quality assurance roles in manufacturing'),
        ('Implementation Timeline: 4-5 months', 'Based on industry average for QMS + EBR implementations'),
        ('No Ongoing Maintenance Costs', 'SaaS subscription includes all maintenance, updates, support'),
        ('3-Year Analysis Period', 'Standard ROI analysis timeframe for enterprise software'),
    ]

    row = 4
    for assumption, justification in assumptions:
        ws_assumptions[f'A{row}'] = assumption
        ws_assumptions[f'A{row}'].font = Font(bold=True)
        ws_assumptions[f'B{row}'] = justification
        ws_assumptions[f'B{row}'].font = Font(size=9)
        ws_assumptions[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1

    # Note section
    ws_assumptions[f'A{row+2}'] = 'COLOR CODING LEGEND:'
    ws_assumptions[f'A{row+2}'].font = Font(bold=True, size=12)

    ws_assumptions[f'A{row+3}'] = 'Blue cells'
    ws_assumptions[f'A{row+3}'].fill = BLUE_FILL
    ws_assumptions[f'B{row+3}'] = 'User inputs (editable values)'

    ws_assumptions[f'A{row+4}'] = 'Black cells'
    ws_assumptions[f'B{row+4}'] = 'Formulas (calculated values, do not edit)'

    ws_assumptions[f'A{row+5}'] = 'Green text'
    ws_assumptions[f'A{row+5}'].font = GREEN_FONT
    ws_assumptions[f'B{row+5}'] = 'Cross-sheet references (pulls from other tabs)'

    # Save workbook
    wb.save(output_path)
    print(f"✅ ROI template created: {output_path}")
    return output_path

if __name__ == '__main__':
    # Create template in assets directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    assets_dir = os.path.join(os.path.dirname(script_dir), 'assets')
    os.makedirs(assets_dir, exist_ok=True)

    output_file = os.path.join(assets_dir, 'roi_template.xlsx')
    create_roi_template(output_file)
